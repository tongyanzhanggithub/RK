# -*- coding: utf-8 -*-
"""汽摩通用配件·多维综合选品(含重庆供应链)：价差+销量+毛利+重庆供应链+竞争+物流+售后+合规，加权排名。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

NAVY="1F3A5F"; PANEL="F4F6F8"; LINE="C9D2DC"; WIN="D6F5D6"; INPUT_BG="FFF4C2"; WHITE="FFFFFF"; GREY="EFEFEF"; CQBG="FCE4D6"
FONT="Arial"
def f(sz=10,b=False,color="000000"): return Font(name=FONT,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook()

# ============== 使用说明 ==============
ws0=wb.active; ws0.title="使用说明"; ws0.sheet_view.showGridLines=False
ws0.column_dimensions["A"].width=3; ws0.column_dimensions["B"].width=110
def Ln(r,t,sz=10,b=False,c="000000"):
    cell=ws0.cell(row=r,column=2,value=t); cell.font=f(sz,b,c); cell.alignment=left
Ln(2,"汽摩通用配件 · 多维综合选品（含重庆供应链）",16,True,NAVY)
Ln(3,"8维加权：价差×销量×毛利×重庆供应链×竞争×物流×售后×合规。面向 英国/中东/东南亚。",10,False,"555555")
blk=[
 (5,"核心思路：重庆供应链 = 你的护城河",12,True,"C0504D"),
 (6,"网站定位是『重庆工厂直供』——重庆是全球摩托/小型发动机产业之都(Loncin/Zongshen/Lifan)。摩配件你能工厂直供、OEM、低MOQ、成本最低，别人难抄；",10),
 (7,"而 CarPlay盒子/记录仪/OBD 这类深圳电子件你没有供应链优势，只能和深圳卖家拼价。所以选品应优先压在『重庆能造、又有价差和需求』的摩配件上。",10,False,"9A6A00"),
 (9,"八个评估维度（每项 1–5 分，5 最好）",12,True,NAVY),
 (10,"1) 价差评分(自动)：英国零售÷国内批发倍数。≥6→5，4.5–6→4，3.5–4.5→3，2.5–3.5→2，<2.5→1。",10),
 (11,"2) 销量需求：海外搜索/成交热度。5=爆款刚需。【手填】",10),
 (12,"3) 毛利评分(自动)：单件绝对赚头$。≥40→5，≥25→4，≥15→3，≥8→2，<8→1。",10),
 (13,"4) 重庆供应链匹配：5=重庆核心集群可工厂直供OEM(摩托CNC件/转向灯/喇叭)；3=部分可配；1=深圳电子件、无本地优势。【手填，护城河维度】",10,False,"C0504D"),
 (14,"5) 竞争蓝海：5=卖家少利润厚；1=红海拼价(通用手机支架/LED灯条已红海)。【手填】",10),
 (15,"6) 物流友好：5=轻小耐摔；1=笨重/易碎/带电受限(充气泵/启动电源运费高)。【手填】",10),
 (16,"7) 售后稳：5=结构简单退货少；1=固件/兼容问题多(OBD/CarPlay/记录仪退货高)。【手填】",10),
 (17,"8) 合规畅：5=无门槛；1=有法规障碍(LED大灯过不了英国MOT、喇叭分贝限制、锂电运输受限)。【手填】",10),
 (19,"权重与排名",12,True,NAVY),
 (20,"顶部第5行是8项权重(蓝色可改，合计须=100%)。默认：价差18% 销量20% 毛利12% 重庆供应链20% 竞争10% 物流8% 售后6% 合规6%。",10),
 (21,"综合评分=各维度分×权重之和，按综合评分排名，前10自动『✅入选』整行绿底。想更激进押供应链护城河，把『重庆供应链』权重调到30%+即可。",10),
 (23,"区域 & 合规提醒",12,True,NAVY),
 (24,"· 东南亚(摩托海量)：摩托转向灯/喇叭/CNC手把/大灯最旺——与重庆供应链高度重合，主战场。· 中东：氛围灯/灯条/记录仪/CarPlay。· 英国：OBD/CarPlay/记录仪。",10),
 (25,"· 英国：LED大灯不过MOT、喇叭有分贝限制→标注 off-road/show use。电子件英UKCA/CE、中东SASO/G-mark，量大再备证。海外价为零售天花板，国内批发为估值，黄格子可覆盖、排名自动刷新。",10,False,"9A6A00"),
]
for x in blk: Ln(*x)

# ============== 综合评估 ==============
ws=wb.create_sheet("综合评估"); ws.sheet_view.showGridLines=False
RATE=7.2
ws.cell(row=2,column=2,value="汇率 ¥/$").font=f(9,False,"555555")
c=ws.cell(row=2,column=3,value=RATE); c.font=f(9,True,"0000FF"); c.fill=fill(INPUT_BG); c.number_format='0.00'; c.alignment=center; c.border=border

cols=[("#",4),("配件(中)",21),("Product (EN)",22),("类型",7),
 ("批发¥低",8),("批发¥高",8),("批发$中",8),
 ("英国$",7),("东南亚$",8),("中东$",7),
 ("价差倍数",8),("毛利$",7),
 ("价差分\n(自动)",7),("销量\n需求",6),("毛利分\n(自动)",7),("重庆\n供应链",7),("竞争\n蓝海",6),("物流\n友好",6),("售后\n稳",5),("合规\n畅",5),
 ("综合\n评分",7),("排名",5),("建议",12)]
WROW=5; HROW=6
for j,(h,w) in enumerate(cols,1):
    ws.column_dimensions[get_column_letter(j)].width=w
# 权重行：8项权重对齐 M..T = 13..20
ws.cell(row=WROW,column=12,value="权重→").font=f(9,True,"555555")
ws.cell(row=WROW,column=12).alignment=Alignment(horizontal="right",vertical="center")
# (col, weight)  M价差 N销量 O毛利 P重庆 Q竞争 R物流 S售后 T合规
wdef=[(13,0.18),(14,0.20),(15,0.12),(16,0.20),(17,0.10),(18,0.08),(19,0.06),(20,0.06)]
for col,val in wdef:
    cc=ws.cell(row=WROW,column=col,value=val); cc.font=f(9,True,"0000FF"); cc.fill=fill(INPUT_BG); cc.number_format='0%'; cc.alignment=center; cc.border=border
cc=ws.cell(row=WROW,column=21,value="=SUM(M5:T5)"); cc.font=f(9,True); cc.number_format='0%'; cc.alignment=center; cc.border=border
ws.cell(row=WROW,column=22,value='=IF(ROUND(U5,4)=1,"✓","≠100%")').font=f(9,True,"C00000")
ws.cell(row=WROW,column=22).alignment=center

for j,(h,w) in enumerate(cols,1):
    cc=ws.cell(row=HROW,column=j,value=h); cc.font=f(8,True,WHITE)
    cc.fill=fill("C0504D") if j==16 else fill(NAVY)
    cc.alignment=center; cc.border=border
ws.row_dimensions[HROW].height=30

# 数据: lo,hi,uk,sea,me, demand,comp,ship,after,compliance,cq
data=[
 ("OBD2诊断仪 ELM327","ELM327 OBD2 scanner","通用",12,30,25,10,13,4,2,5,2,4,1),
 ("H4/H7 LED大灯(对)","H4/H7 LED headlight","车/摩",40,90,50,26,30,5,2,5,3,2,3),
 ("摩托LED转向灯(对)","Moto LED indicators","摩托",15,40,20,10,13,4,3,5,4,3,5),
 ("无线CarPlay/Auto盒子","Wireless CarPlay adapter","车",90,150,75,50,55,5,3,5,2,5,1),
 ("CNC摩托手把/镜/牛角","CNC moto grips/mirrors","摩托",35,90,40,22,26,3,4,4,5,5,5),
 ("车内氛围灯RGB(App)","Ambient RGB light kit","车",30,65,30,18,22,4,3,5,4,5,2),
 ("蜗牛/气喇叭","Snail/air horn","通用",20,50,22,11,14,3,4,4,5,3,4),
 ("LED射灯/灯条","LED spotlight/bar","通用",45,110,48,30,35,4,3,4,4,3,3),
 ("磁吸/无线手机支架","Phone holder","通用",15,40,16,8,10,5,2,5,5,5,3),
 ("行车记录仪1080p/2K","Dash cam","车",60,130,55,40,45,5,2,4,3,4,1),
 ("便携数显充气泵","Tyre inflator","车",70,130,42,30,38,5,3,3,3,5,2),
 ("太阳能TPMS胎压","Solar TPMS","车",60,110,38,25,30,3,3,5,3,5,1),
 ("无线倒车影像","Wireless rear camera","车",80,150,60,40,45,3,3,4,3,5,1),
 ("车载无线充支架","Wireless charge mount","车",55,110,42,28,32,4,3,5,4,5,1),
 ("应急启动电源","Jump starter","车",100,160,45,38,42,4,2,2,3,3,1),
]
first=HROW+1; n=len(data); last=first+n-1
for i,row in enumerate(data):
    cn,en,tp,lo,hi,uk,sea,me,dem,comp,ship,aft,cpl,cq=row
    r=first+i
    def cell(col,val,nf=None,bold=False,color="000000",inp=False,al=center,auto=False,cqf=False):
        cc=ws.cell(row=r,column=col,value=val)
        cc.font=f(8,bold,("0000FF" if (inp or cqf) else color)); cc.alignment=al; cc.border=border
        if nf: cc.number_format=nf
        if cqf: cc.fill=fill(CQBG)
        elif inp: cc.fill=fill(INPUT_BG)
        elif auto: cc.fill=fill(GREY)
        return cc
    cell(1,i+1); cell(2,cn,al=left); cell(3,en,al=left); cell(4,tp)
    cell(5,lo,'¥#,##0',inp=True); cell(6,hi,'¥#,##0',inp=True)
    cell(7,f"=(E{r}+F{r})/2/$C$2",'$#,##0.0')
    cell(8,uk,'$#,##0',inp=True); cell(9,sea,'$#,##0',inp=True); cell(10,me,'$#,##0',inp=True)
    cell(11,f"=IF(G{r}=0,\"\",H{r}/G{r})",'0.0"x"',bold=True,color=NAVY)
    cell(12,f"=H{r}-G{r}",'$#,##0.0')
    cell(13,f'=IF(K{r}="","",IF(K{r}>=6,5,IF(K{r}>=4.5,4,IF(K{r}>=3.5,3,IF(K{r}>=2.5,2,1)))))',auto=True)
    cell(14,dem,inp=True)
    cell(15,f'=IF(L{r}="","",IF(L{r}>=40,5,IF(L{r}>=25,4,IF(L{r}>=15,3,IF(L{r}>=8,2,1)))))',auto=True)
    cell(16,cq,cqf=True,bold=True)   # 重庆供应链
    cell(17,comp,inp=True); cell(18,ship,inp=True); cell(19,aft,inp=True); cell(20,cpl,inp=True)
    cell(21,f"=M{r}*$M$5+N{r}*$N$5+O{r}*$O$5+P{r}*$P$5+Q{r}*$Q$5+R{r}*$R$5+S{r}*$S$5+T{r}*$T$5",'0.00',bold=True)
    cell(22,f"=RANK(U{r},$U${first}:$U${last})",bold=True)
    cell(23,f'=IF(V{r}<=10,"✅ 入选 Top10","备选")',bold=True)

ws.freeze_panes="E7"
ws.conditional_formatting.add(f"A{first}:W{last}",
    FormulaRule(formula=[f"$V{first}<=10"], fill=fill(WIN)))
nr=last+2
ws.cell(row=nr,column=2,value="黄=可现场覆盖输入；橙=重庆供应链匹配(护城河维度)；灰=自动评分。按综合评分排名，前10绿底✅。改权重(第5行)或任意输入，排名自动刷新。")
ws.cell(row=nr,column=2).font=f(9,False,"9A6A00")

wb.save("汽摩通用配件-综合选品.xlsx")
print("saved rows",first,"-",last)
