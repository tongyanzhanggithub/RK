# -*- coding: utf-8 -*-
"""汽摩通用配件·国内外价差选品表：找国内批发便宜、海外(英国/中东/东南亚)零售贵、价差倍数最大的件。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3A5F"; PANEL="F4F6F8"; LINE="C9D2DC"; WIN="D6F5D6"; INPUT_BG="FFF4C2"; WHITE="FFFFFF"
FONT="Arial"
def f(sz=10,b=False,color="000000"): return Font(name=FONT,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook()

# ---------- 使用说明 ----------
ws0=wb.active; ws0.title="使用说明"; ws0.sheet_view.showGridLines=False
ws0.column_dimensions["A"].width=3; ws0.column_dimensions["B"].width=104
def L(r,t,sz=10,b=False,c="000000"):
    cell=ws0.cell(row=r,column=2,value=t); cell.font=f(sz,b,c); cell.alignment=left
L(2,"汽摩通用配件 · 国内外价差选品",16,True,NAVY)
L(3,"目标：在汽车/摩托通用配件里，挑『国内批发极便宜、海外零售很贵』价差倍数最大的件，面向 英国 / 中东 / 东南亚",10,False,"555555")
rows=[
 (5,"结论速览",12,True,NAVY),
 (6,"价差倍数最大的集中在『电子类 + 改装类通用件』：OBD诊断、LED大灯/转向灯/灯条、无线CarPlay盒子、氛围灯、喇叭、手机支架等。",10),
 (7,"这些件国内批发常 ¥10–100，海外零售 $20–80，价差 4–9 倍；且体积小、重量轻、跨车型通用，最适合独立站走量试销。",10),
 (9,"表格怎么看（『价差排行』标签页）",12,True,NAVY),
 (10,"· 国内批发价：1688/AliExpress 批发档低-高区间(¥)，表格自动取中位换成美元。",10),
 (11,"· 英国/中东/东南亚零售价($)：Amazon UK / noon / Lazada-Shopee 同类件零售挂牌价。",10),
 (12,"· 价差倍数 = 英国零售$ ÷ 国内批发$(中)。这是核心排序指标，倍数越高越值得做。",10),
 (13,"· 单件毛利≈ 英国零售$ − 国内批发$(中)，看绝对赚头(未扣运费/平台费，仅作横向对比)。",10),
 (14,"· 建议列：价差倍数排名前 10 自动标『✅ 入选』(绿底)，其余为『备选』。",10),
 (16,"区域匹配提示",12,True,NAVY),
 (17,"· 东南亚(印尼/越南/泰国/菲律宾)：摩托车海量 → 摩托LED转向灯、喇叭、CNC手把/后视镜、大灯最旺。",10),
 (18,"· 中东(沙特/阿联酋)：汽车改装+高温 → 氛围灯、灯条、行车记录仪、充气泵、CarPlay 盒子需求强。",10),
 (19,"· 英国：客单价高但合规严 → OBD/CarPlay/记录仪/充气泵好卖；但 LED大灯改装不过 MOT 年检、喇叭有分贝限制，UK 站点需标注『off-road / show use』。",10,False,"9A6A00"),
 (21,"重要提醒",12,True,NAVY),
 (22,"1) 海外价为零售挂牌价(2026-06)，代表价格天花板；你的独立站定价可略低于它抢量，不要直接照搬。",10),
 (23,"2) 国内批发价为常见批发档估值，黄色格子可按你市场实拿价覆盖，倍数与排名会自动刷新。",10,False,"9A6A00"),
 (24,"3) 电子类(OBD/记录仪/启动电源)注意各国认证：英国 UKCA/CE、中东 SASO/G-mark，量大再备证。",10,False,"9A6A00"),
]
for r in rows:
    L(*r)

# ---------- 价差排行 ----------
ws=wb.create_sheet("价差排行"); ws.sheet_view.showGridLines=False
RATE=7.2
ws["B2"]="汇率 ¥/$"; ws["B2"].font=f(9,False,"555555")
c=ws["C2"]; c.value=RATE; c.font=f(9,True,"0000FF"); c.fill=fill(INPUT_BG); c.number_format='0.00'; c.alignment=center; c.border=border

HROW=4
cols=[("#",4),("配件(中)",24),("Product (EN)",26),("类型",8),
 ("国内批发¥低",10),("国内批发¥高",10),("国内批发$中",11),
 ("英国零售$",10),("东南亚零售$",11),("中东零售$",10),
 ("价差倍数(UK/CN)",12),("单件毛利$≈",11),("物流轻小1-5",9),
 ("最旺区域",12),("建议",13)]
for j,(h,w) in enumerate(cols,1):
    cc=ws.cell(row=HROW,column=j,value=h); cc.font=f(9,True,WHITE); cc.fill=fill(NAVY)
    cc.alignment=center; cc.border=border; ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[HROW].height=32

# data: name_cn, en, type, cn_low, cn_high, uk, sea, me, ship, region
data=[
 ("OBD2诊断仪 ELM327(蓝牙)","ELM327 OBD2 scanner (BT)","通用车",12,30,25,10,13,5,"英国/中东"),
 ("H4/H7 LED大灯(对)","H4/H7 LED headlight kit","车/摩",40,90,50,26,30,5,"东南亚/中东"),
 ("摩托LED转向灯(对)","Motorcycle LED indicators","摩托",15,40,20,10,13,5,"东南亚"),
 ("无线CarPlay/安卓Auto盒子","Wireless CarPlay/AA adapter","车",90,150,75,50,55,5,"中东/英国"),
 ("CNC摩托手把/后视镜/牛角","CNC moto grips/mirrors/levers","摩托",35,90,40,22,26,4,"东南亚"),
 ("车内氛围灯RGB(App控)","Car ambient RGB light kit","车",30,65,30,18,22,5,"中东"),
 ("蜗牛喇叭/气喇叭","Snail/air horn","通用",20,50,22,11,14,4,"东南亚/中东"),
 ("LED射灯/灯条(摩/越野)","LED spotlight/light bar","通用",45,110,48,30,35,4,"中东/东南亚"),
 ("磁吸/无线手机支架","Magnetic/wireless phone holder","通用",15,40,16,8,10,5,"东南亚"),
 ("行车记录仪 1080p/2K","Dash cam 1080p/2K","车",60,130,55,40,45,5,"中东/英国"),
 ("便携数显充气泵","Portable digital tyre inflator","车",70,130,42,30,38,4,"中东/英国"),
 ("太阳能胎压监测 TPMS","Solar TPMS","车",60,110,38,25,30,5,"中东"),
 ("无线倒车影像套装","Wireless reversing camera","车",80,150,60,40,45,4,"中东"),
 ("车载无线充支架","Wireless charging mount","车",55,110,42,28,32,5,"中东/英国"),
 ("应急启动电源","Jump starter power bank","车",100,160,45,38,42,3,"英国/中东"),
]
first=HROW+1; n=len(data); last=first+n-1
for i,(cn,en,tp,lo,hi,uk,sea,me,ship,reg) in enumerate(data):
    r=first+i
    def put(col,val,nf=None,bold=False,color="000000",inp=False,al=center):
        cell=ws.cell(row=r,column=col,value=val); cell.font=f(9,bold,color); cell.alignment=al; cell.border=border
        if nf: cell.number_format=nf
        if inp: cell.fill=fill(INPUT_BG); cell.font=f(9,bold,"0000FF")
    put(1,i+1)
    put(2,cn,al=left)
    put(3,en,al=left)
    put(4,tp)
    put(5,lo,'¥#,##0',inp=True)
    put(6,hi,'¥#,##0',inp=True)
    ws.cell(row=r,column=7,value=f"=(E{r}+F{r})/2/$C$2");
    cc=ws.cell(row=r,column=7); cc.font=f(9); cc.number_format='$#,##0.00'; cc.alignment=center; cc.border=border
    put(8,uk,'$#,##0',inp=True)
    put(9,sea,'$#,##0',inp=True)
    put(10,me,'$#,##0',inp=True)
    # 价差倍数
    cc=ws.cell(row=r,column=11,value=f"=IF(G{r}=0,\"\",H{r}/G{r})"); cc.font=f(10,True,NAVY); cc.number_format='0.0"x"'; cc.alignment=center; cc.border=border
    # 毛利
    cc=ws.cell(row=r,column=12,value=f"=H{r}-G{r}"); cc.font=f(9); cc.number_format='$#,##0.0'; cc.alignment=center; cc.border=border
    put(13,ship)
    put(14,reg,al=center)
    # 建议
    cc=ws.cell(row=r,column=15,value=f'=IF(RANK(K{r},$K${first}:$K${last})<=10,"✅ 入选 Top10","备选")'); cc.font=f(9,True); cc.alignment=center; cc.border=border

ws.freeze_panes="A5"
# 条件格式：价差前10整行绿底
from openpyxl.formatting.rule import FormulaRule
rng=f"A{first}:O{last}"
ws.conditional_formatting.add(rng,
    FormulaRule(formula=[f"RANK($K{first},$K${first}:$K${last})<=10"], fill=fill(WIN)))
# 加排名列说明在表下
nr=last+2
ws.cell(row=nr,column=2,value="排序按『价差倍数』；黄色格子(国内批发价/海外零售价)可按你实拿价覆盖，倍数与建议自动刷新。绿底=价差前10。")
ws.cell(row=nr,column=2).font=f(9,False,"9A6A00")

wb.save("汽摩通用配件-价差选品.xlsx")
print("saved rows",first,"-",last)
