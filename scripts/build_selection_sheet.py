# -*- coding: utf-8 -*-
"""生成『选品打分表』：高频刚需走量导向，对比国内进货价 vs 海外售价。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

NAVY = "1F3A5F"
SAFETY = "F2B705"
PANEL = "F4F6F8"
LINE = "C9D2DC"
INPUT_BG = "FFF4C2"   # 需现场填写
WHITE = "FFFFFF"

FONT = "Arial"
def f(sz=10, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)
def fill(c):
    return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ================= Sheet 1: 说明 =================
ws0 = wb.active
ws0.title = "使用说明"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 3
ws0.column_dimensions["B"].width = 100

def line(ws, r, text, sz=10, b=False, color="000000"):
    c = ws.cell(row=r, column=2, value=text)
    c.font = f(sz, b, color); c.alignment = left

ws0["B2"] = "汽配独立站 · 选品打分表"
ws0["B2"].font = f(16, True, NAVY)
ws0["B3"] = "目标：高频刚需走量 —— 优先选维修店天天要换、适配机型多、海外有售价空间的易损件"
ws0["B3"].font = f(10, False, "555555")

rows = [
    (5,  "怎么用（3 步）", 12, True, NAVY),
    (6,  "1. 去市场前先看『选品打分表』里我预填的 15 个候选件 —— 都是 168F/GX160/GX200/170F/水泵/发电机的高频易损件。", 10, False, "000000"),
    (7,  "2. 现场把【黄色格子】填上：国内进货价(¥)、起订量、单件重量(g)，再按 1-5 给『更换频率/通用性/竞争激烈度』打分。", 10, False, "000000"),
    (8,  "3. 表格自动算出毛利率和综合评分并排名，『建议』列里标 ✅ 的就是排名前 10、可上架试销的。", 10, False, "000000"),
    (10, "黄色格子 = 必须现场填 / 可改", 10, True, "9A6A00"),
    (11, "蓝字 = 假设值（汇率、各项权重），按你实际情况改；黑字 = 公式自动算，别动。", 10, False, "555555"),
    (13, "评分口径（1-5 分，5 最好）", 12, True, NAVY),
    (14, "· 更换频率/刚需度：维修店多久换一次。5=几乎每台都换(化油器/火花塞/密封)，1=很少坏。", 10, False, "000000"),
    (15, "· 通用性：能装多少种机型。5=几乎所有168F/GX系列通用，1=只配单一型号。机型越多越好走量、库存越简单。", 10, False, "000000"),
    (16, "· 价格竞争力：表格按『毛利率』自动评分，不用手填。", 10, False, "000000"),
    (17, "· 物流友好：表格按『重量』自动评分(越轻小越好)，不用手填。", 10, False, "000000"),
    (18, "· 竞争激烈度(蓝海度)：5=海外卖家少/差异化好，1=红海拼价。需自己判断填。", 10, False, "000000"),
    (20, "关于『海外售价参考』", 12, True, NAVY),
    (21, "数据来自 Amazon / eBay / Walmart / AliExpress 同类售后件实时挂牌价(2026-06)，取的是售后件(非原厂OEM)价格区间。", 10, False, "000000"),
    (22, "它代表海外零售『需求 + 价格天花板』，不是你的批发售价；毛利率算的是『相对海外零售的空间』，用来横向比哪个件更值得做。", 10, False, "555555"),
    (23, "明细见『海外售价参考』标签页。", 10, False, "555555"),
]
for r, t, sz, b, c in rows:
    line(ws0, r, t, sz, b, c)

# ================= Sheet 3 data: 海外售价参考 =================
# (name_cn, name_en, cat, lo, hi, typ, src)
ref = [
    ("化油器", "Carburetor (GX160/GX200/168F)", "燃油系统", 10, 18, 14, "Amazon 售后件 HIPA/Savior/Anxingo"),
    ("化油器修理包", "Carb rebuild kit (needle/float/gasket)", "燃油系统", 7, 12, 9, "Amazon Sellerocity/通用修理包"),
    ("手拉启动器总成", "Recoil starter assembly", "启动系统", 15, 42, 22, "Amazon FDJ/Stens/OAKTEN 售后件"),
    ("点火线圈", "Ignition coil", "点火系统", 10, 25, 16, "Amazon Atunee/通用件"),
    ("水泵机械密封", "Water pump mechanical seal", "水泵", 7, 12, 9, "AliExpress/Amazon 78130-YB4 类"),
    ("水泵全套胶圈密封包", "Pump full seal/gasket kit", "水泵", 12, 20, 15, "Amazon/wsepo 2寸泵密封包"),
    ("发电机 AVR", "Generator AVR (2-3kW)", "发电机", 9, 15, 12, "Amazon/Walmart 168F 2-3kW AVR"),
    ("活塞环+大修垫片包", "Piston ring + gasket rebuild kit", "发动机大修", 15, 24, 19, "Amazon HAISHINE/ApplianPar; eBay"),
    ("全套垫片", "Full gasket set", "发动机", 10, 18, 13, "Amazon TXGXMB 垫片套"),
    ("火花塞", "Spark plug", "点火系统", 3, 8, 5, "Amazon 通用件(常以套装出现)"),
    ("空气滤芯", "Air filter", "进气", 4, 9, 6, "Amazon 通用件"),
    ("燃油开关/油杯", "Fuel petcock / tap", "燃油系统", 5, 10, 7, "Amazon/eBay 通用件"),
    ("启动绳+手柄", "Pull rope + handle", "启动系统", 4, 8, 6, "Amazon 通用件"),
    ("发电机碳刷", "Carbon brush set", "发电机", 5, 10, 7, "Amazon/eBay 通用件"),
    ("调试套装(化油器+空滤+火花塞)", "Tune-up combo kit", "组合套件", 15, 25, 19, "Amazon 组合套装"),
]

ws3 = wb.create_sheet("海外售价参考")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "海外售价参考（USD · 售后件挂牌价 · 2026-06）"
ws3["A1"].font = f(14, True, NAVY)
ws3["A2"] = "来源：Amazon / eBay / Walmart / AliExpress。代表海外零售需求与价格天花板，非批发售价。"
ws3["A2"].font = f(9, False, "555555")
hdr3 = ["#", "产品(中)", "Product (EN)", "品类", "最低价 $", "最高价 $", "参考价 $(典型)", "数据来源"]
for j, h in enumerate(hdr3, 1):
    c = ws3.cell(row=4, column=j, value=h)
    c.font = f(10, True, WHITE); c.fill = fill(NAVY); c.alignment = center; c.border = border
for i, (cn, en, cat, lo, hi, typ, src) in enumerate(ref):
    r = 5 + i
    vals = [i+1, cn, en, cat, lo, hi, typ, src]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=r, column=j, value=v)
        c.font = f(9); c.border = border
        c.alignment = center if j in (1,4,5,6,7) else left
        if j in (5,6,7): c.number_format = '$#,##0.00'
        if r % 2 == 0: c.fill = fill(PANEL)
    ws3.cell(row=r, column=7).font = f(9, True, NAVY)
widths3 = [4, 26, 34, 12, 10, 10, 13, 34]
for j, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.freeze_panes = "A5"

# ================= Sheet 2: 选品打分表 =================
ws = wb.create_sheet("选品打分表")
wb.move_sheet("选品打分表", -(len(wb.sheetnames)-1))  # put after 使用说明
ws.sheet_view.showGridLines = False

# --- assumptions block ---
ws["B2"] = "假设（蓝字可改）"
ws["B2"].font = f(11, True, NAVY)
asm = [
    ("汇率 ¥/$", 7.20, "B3", "C3", '0.00'),
    ("权重·更换频率",  0.35, "B4", "C4", '0.0%'),
    ("权重·通用性",    0.25, "B5", "C5", '0.0%'),
    ("权重·价格竞争力", 0.20, "B6", "C6", '0.0%'),
    ("权重·物流友好",   0.10, "B7", "C7", '0.0%'),
    ("权重·竞争蓝海",   0.10, "B8", "C8", '0.0%'),
]
for label, val, lc, vc, nf in asm:
    ws[lc] = label; ws[lc].font = f(9, False, "555555"); ws[lc].alignment = left
    cell = ws[vc]; cell.value = val; cell.font = f(9, True, "0000FF")
    cell.fill = fill(INPUT_BG); cell.number_format = nf; cell.alignment = center; cell.border = border
ws["B9"] = "权重合计"; ws["B9"].font = f(9, False, "555555")
ws["C9"] = "=SUM(C4:C8)"; ws["C9"].font = f(9, True); ws["C9"].number_format = '0.0%'
ws["C9"].alignment = center; ws["C9"].border = border
ws["D9"] = '=IF(ROUND(C9,4)=1,"OK","权重之和必须=100%")'
ws["D9"].font = f(9, True, "C00000"); ws["D9"].alignment = left

# --- table header (row 11) ---
HROW = 11
cols = [
    ("#", 4),
    ("产品(候选件)", 28),
    ("品类", 11),
    ("海外参考价 $", 12),
    ("国内进货价 ¥", 13),
    ("进货价 $", 10),
    ("毛利额 $", 10),
    ("毛利率", 9),
    ("起订量 MOQ", 11),
    ("单件重量 g", 11),
    ("更换频率 1-5", 10),
    ("通用性 1-5", 10),
    ("竞争蓝海 1-5", 10),
    ("价格力(自动)", 10),
    ("物流力(自动)", 10),
    ("综合评分", 10),
    ("排名", 7),
    ("建议", 14),
]
for j, (h, w) in enumerate(cols, 1):
    c = ws.cell(row=HROW, column=j, value=h)
    c.font = f(9, True, WHITE); c.fill = fill(NAVY); c.alignment = center; c.border = border
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[HROW].height = 34

# mark input headers with safety underline note via comment
for col, note in [(5,"现场填：国内进货价(人民币)"),(9,"现场填：最小起订量"),(10,"现场填：单件净重(克)"),
                  (11,"打分1-5：维修店换得多不多"),(12,"打分1-5：适配机型多不多"),(13,"打分1-5：5=蓝海少竞争")]:
    ws.cell(row=HROW, column=col).comment = Comment(note, "选品表")

# --- data rows ---
# pre-filled suggestions: freq, univ, weight, moq(blank), competition (blank-ish)
seed = [
    # name, cat, refp, freq, univ, weight_g
    ("化油器 Carburetor",            "燃油系统", 14, 5, 5, 350),
    ("化油器修理包 Carb kit",         "燃油系统",  9, 5, 4, 50),
    ("手拉启动器总成 Recoil starter", "启动系统", 22, 4, 5, 500),
    ("点火线圈 Ignition coil",        "点火系统", 16, 4, 5, 200),
    ("水泵机械密封 Pump seal",        "水泵",     9, 5, 4, 60),
    ("水泵全套胶圈包 Pump seal kit",  "水泵",    15, 4, 3, 150),
    ("发电机 AVR",                    "发电机",  12, 3, 3, 120),
    ("活塞环+大修垫片包 Rebuild kit", "大修",    19, 3, 4, 400),
    ("全套垫片 Gasket set",           "发动机",  13, 4, 4, 80),
    ("火花塞 Spark plug",             "点火系统",  5, 5, 5, 50),
    ("空气滤芯 Air filter",           "进气",     6, 5, 5, 40),
    ("燃油开关/油杯 Fuel petcock",    "燃油系统",  7, 4, 4, 60),
    ("启动绳+手柄 Pull rope",         "启动系统",  6, 4, 5, 50),
    ("发电机碳刷 Carbon brush",       "发电机",   7, 3, 3, 30),
    ("调试套装 Tune-up combo",        "组合套件", 19, 4, 4, 450),
]
first = HROW + 1
n = len(seed)
last = first + n - 1
for i, (name, cat, refp, freq, univ, wt) in enumerate(seed):
    r = first + i
    # 1 #
    ws.cell(row=r, column=1, value=i+1).font = f(9)
    ws.cell(row=r, column=1).alignment = center
    # 2 name
    ws.cell(row=r, column=2, value=name).font = f(9); ws.cell(row=r, column=2).alignment = left
    # 3 cat
    ws.cell(row=r, column=3, value=cat).font = f(9); ws.cell(row=r, column=3).alignment = center
    # 4 overseas ref $  (input-ish, prefilled, blue)
    c = ws.cell(row=r, column=4, value=refp); c.font = f(9, False, "0000FF"); c.number_format='$#,##0.00'; c.alignment=center
    # 5 china cost RMB (INPUT yellow)
    c = ws.cell(row=r, column=5); c.fill = fill(INPUT_BG); c.font = f(9, False, "0000FF"); c.number_format='¥#,##0.00'; c.alignment=center
    # 6 cost $ = rmb / rate
    c = ws.cell(row=r, column=6, value=f"=IF(E{r}=\"\",\"\",E{r}/$C$3)"); c.font=f(9); c.number_format='$#,##0.00'; c.alignment=center
    # 7 margin $ = ref - cost$
    c = ws.cell(row=r, column=7, value=f"=IF(F{r}=\"\",\"\",D{r}-F{r})"); c.font=f(9); c.number_format='$#,##0.00'; c.alignment=center
    # 8 margin %
    c = ws.cell(row=r, column=8, value=f"=IF(OR(F{r}=\"\",D{r}=0),\"\",G{r}/D{r})"); c.font=f(9); c.number_format='0.0%'; c.alignment=center
    # 9 MOQ (input)
    c = ws.cell(row=r, column=9); c.fill = fill(INPUT_BG); c.font=f(9, False, "0000FF"); c.number_format='#,##0'; c.alignment=center
    # 10 weight (prefilled input)
    c = ws.cell(row=r, column=10, value=wt); c.fill = fill(INPUT_BG); c.font=f(9, False, "0000FF"); c.number_format='#,##0'; c.alignment=center
    # 11 freq (prefilled input)
    c = ws.cell(row=r, column=11, value=freq); c.fill = fill(INPUT_BG); c.font=f(9, False, "0000FF"); c.alignment=center
    # 12 univ (prefilled input)
    c = ws.cell(row=r, column=12, value=univ); c.fill = fill(INPUT_BG); c.font=f(9, False, "0000FF"); c.alignment=center
    # 13 competition blue ocean (input, default 3)
    c = ws.cell(row=r, column=13, value=3); c.fill = fill(INPUT_BG); c.font=f(9, False, "0000FF"); c.alignment=center
    # 14 price score auto from margin%
    c = ws.cell(row=r, column=14,
        value=(f'=IF(H{r}="","",IF(H{r}>=0.65,5,IF(H{r}>=0.5,4,IF(H{r}>=0.35,3,IF(H{r}>=0.2,2,1)))))'))
    c.font=f(9); c.alignment=center
    # 15 logistics score auto from weight
    c = ws.cell(row=r, column=15,
        value=(f'=IF(J{r}="","",IF(J{r}<=100,5,IF(J{r}<=300,4,IF(J{r}<=600,3,IF(J{r}<=1000,2,1)))))'))
    c.font=f(9); c.alignment=center
    # 16 composite score
    c = ws.cell(row=r, column=16,
        value=(f'=IF(OR(K{r}="",L{r}="",M{r}="",N{r}="",O{r}=""),"",'
               f'K{r}*$C$4+L{r}*$C$5+N{r}*$C$6+O{r}*$C$7+M{r}*$C$8)'))
    c.font=f(9, True); c.number_format='0.00'; c.alignment=center
    # 17 rank
    c = ws.cell(row=r, column=17, value=f'=IF(P{r}="","",RANK(P{r},$P${first}:$P${last}))')
    c.font=f(9, True); c.alignment=center
    # 18 recommendation
    c = ws.cell(row=r, column=18, value=f'=IF(Q{r}="","",IF(Q{r}<=10,"✅ 入选试销","观望"))')
    c.font=f(9, True); c.alignment=center
    for j in range(1, 19):
        ws.cell(row=r, column=j).border = border

ws.freeze_panes = "A12"

# conditional-ish color for 建议 via formula font cannot; leave plain. Add note row
note_r = last + 2
ws.cell(row=note_r, column=2,
    value="提示：黄色列现场填(国内进货价/MOQ/重量/三项评分)；进货价、毛利、评分、排名、建议自动算。排名前10=建议上架试销。")
ws.cell(row=note_r, column=2).font = f(9, False, "9A6A00")

wb.save("选品打分表.xlsx")
print("saved 选品打分表.xlsx rows", first, "-", last)
